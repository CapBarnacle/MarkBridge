"""Basic parser implementations for currently enabled formats."""

from __future__ import annotations

from functools import lru_cache
from pathlib import Path
import re
from tempfile import TemporaryDirectory

from docx import Document as DocxDocument
from docx.document import Document as DocxPackageDocument
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader

from markbridge.parsers.base import ParseRequest, ParseResult
from markbridge.parsers.conversion import (
    TextExtractionResult,
    convert_doc_to_docx,
    extract_doc_text_with_antiword,
    extract_hwp_text_with_hwp5txt,
)
from markbridge.shared.ir import BlockIR, BlockKind, DocumentFormat, DocumentIR, TableBlockIR, TableCellIR


def parse_with_current_runtime(request: ParseRequest, parser_id: str) -> ParseResult:
    if parser_id == "docling":
        return _parse_pdf_with_docling(request)
    if parser_id == "markitdown":
        return _parse_with_markitdown(request)
    if parser_id == "pdfplumber":
        return _parse_pdf_with_pdfplumber(request)
    if parser_id == "pypdf":
        return _parse_pdf(request)
    if parser_id == "python-docx":
        return _parse_docx(request)
    if parser_id == "openpyxl":
        return _parse_xlsx(request)
    if parser_id == "libreoffice":
        return _parse_doc_via_conversion(request)
    if parser_id == "antiword":
        return _parse_doc_with_antiword(request)
    if parser_id == "hwp5txt":
        return _parse_hwp_with_hwp5txt(request)
    raise ValueError(f"Unsupported parser route: {parser_id}")


def _parse_pdf(request: ParseRequest) -> ParseResult:
    reader = PdfReader(str(request.source_path))
    blocks: list[BlockIR] = []
    for page_index, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            blocks.append(
                BlockIR(
                    kind=BlockKind.PARAGRAPH,
                    text=text,
                    metadata={"page": page_index},
                )
            )
    document = DocumentIR(
        source_format=DocumentFormat.PDF,
        blocks=tuple(blocks),
        metadata={"page_count": len(reader.pages)},
    )
    return ParseResult(parser_id="pypdf", document=document)


def _parse_pdf_with_pdfplumber(request: ParseRequest) -> ParseResult:
    import pdfplumber

    blocks: list[BlockIR] = []
    page_count = 0
    with pdfplumber.open(str(request.source_path)) as pdf:
        page_count = len(pdf.pages)
        for page_index, page in enumerate(pdf.pages, start=1):
            text = (page.extract_text() or "").strip()
            if text:
                blocks.append(
                    BlockIR(
                        kind=BlockKind.PARAGRAPH,
                        text=text,
                        metadata={"page": page_index, "source": "pdfplumber.extract_text"},
                    )
                )

    document = DocumentIR(
        source_format=DocumentFormat.PDF,
        blocks=tuple(blocks),
        metadata={"page_count": page_count, "source": "pdfplumber.extract_text"},
    )
    return ParseResult(parser_id="pdfplumber", document=document)


def _parse_pdf_with_docling(request: ParseRequest) -> ParseResult:
    converter = _get_docling_converter()
    conversion = converter.convert(str(request.source_path))
    markdown = conversion.document.export_to_markdown().strip()
    page_count = getattr(request.inspection.common, "page_count", None) if request.inspection else None
    default_page_number = 1 if page_count == 1 else None
    blocks = _blocks_from_markdown(markdown, default_page_number=default_page_number)
    metadata = {
        "preferred_markdown": markdown,
        "source": "docling.export_to_markdown",
        "ocr_disabled": True,
    }
    if page_count is not None:
        metadata["page_count"] = page_count
    document = DocumentIR(
        source_format=DocumentFormat.PDF,
        blocks=tuple(blocks),
        metadata=metadata,
    )
    return ParseResult(parser_id="docling", document=document)


@lru_cache(maxsize=1)
def _get_docling_converter():
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.document_converter import DocumentConverter, PdfFormatOption

    pipeline_options = PdfPipelineOptions(
        do_ocr=False,
        do_picture_classification=False,
        do_picture_description=False,
        do_code_enrichment=False,
        do_formula_enrichment=False,
        generate_page_images=False,
        generate_picture_images=False,
        generate_table_images=False,
    )
    return DocumentConverter(
        allowed_formats=[InputFormat.PDF],
        format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=pipeline_options)},
    )


def _parse_docx(request: ParseRequest) -> ParseResult:
    doc = DocxDocument(str(request.source_path))
    blocks: list[BlockIR] = []
    ordered_items = list(_iter_docx_block_items(doc))
    table_index = 0

    for item_index, item in enumerate(ordered_items):
        if isinstance(item, Paragraph):
            text = item.text.strip()
            if not text:
                continue
            next_item_kind, next_text = _next_nonempty_docx_item_hint(ordered_items, start_index=item_index + 1)
            previous_paragraph_text = _adjacent_nonempty_docx_paragraph_text(
                ordered_items,
                start_index=item_index,
                reverse=True,
            )
            nearby_circled_sequence = _has_nearby_circled_number_paragraph(
                ordered_items,
                start_index=item_index + 1,
            )
            nearby_circled_sequence = nearby_circled_sequence or _has_nearby_circled_number_paragraph(
                ordered_items,
                start_index=item_index,
                reverse=True,
            )
            heading_hint = _docx_heading_hint(
                text=text,
                style_name=item.style.name if item.style else None,
                next_text=next_text,
                next_item_kind=next_item_kind,
                previous_paragraph_text=previous_paragraph_text,
                nearby_circled_sequence=nearby_circled_sequence,
            )
            if heading_hint is not None:
                blocks.append(
                    BlockIR(
                        kind=BlockKind.HEADING,
                        text=text,
                        metadata={
                            "level": heading_hint["level"],
                            "chunk_boundary_candidate": True,
                            "chunk_boundary_reason": heading_hint["reason"],
                            "chunk_boundary_confidence": heading_hint["confidence"],
                            "chunk_boundary_materialized": True,
                        },
                    )
                )
                continue
            blocks.append(BlockIR(kind=BlockKind.PARAGRAPH, text=text))
            continue

        if not isinstance(item, Table):
            continue
        table_index += 1
        layout_rows = _docx_table_rows_text(item)
        if _is_docx_layout_table(layout_rows):
            blocks.extend(_blocks_from_docx_layout_rows(layout_rows))
            continue
        normalized_rows, applied_carry_forward = _normalize_docx_table_rows(item)
        cells: list[TableCellIR] = []
        for row_idx, row_values in enumerate(normalized_rows):
            for col_idx, text in enumerate(row_values):
                if text:
                    cells.append(TableCellIR(row_index=row_idx, column_index=col_idx, text=text, is_header=row_idx == 0))
        if cells:
            blocks.append(
                TableBlockIR(
                    cells=tuple(cells),
                    table_id=f"docx-table-{table_index}",
                    title=f"Table {table_index}",
                    merged_cells=applied_carry_forward,
                    metadata={"docx_carry_forward": applied_carry_forward},
                )
            )

    document = DocumentIR(source_format=DocumentFormat.DOCX, blocks=tuple(blocks), metadata={"table_count": len(doc.tables)})
    return ParseResult(parser_id="python-docx", document=document)


def _parse_with_markitdown(request: ParseRequest) -> ParseResult:
    from markitdown import MarkItDown

    converter = MarkItDown(enable_plugins=False)
    conversion = converter.convert(request.source_path)
    markdown = (conversion.markdown or conversion.text_content or "").strip()
    blocks = _blocks_from_markdown(markdown)
    metadata = {
        "preferred_markdown": markdown,
        "source": "markitdown.convert",
    }
    if conversion.title:
        metadata["title"] = conversion.title
    return ParseResult(
        parser_id="markitdown",
        document=DocumentIR(
            source_format=request.document_format,
            blocks=tuple(blocks),
            metadata=metadata,
        ),
    )


def _parse_xlsx(request: ParseRequest) -> ParseResult:
    workbook = load_workbook(str(request.source_path), data_only=False)
    blocks: list[BlockIR] = []
    table_index = 0
    for sheet in workbook.worksheets:
        if sheet.title.strip():
            blocks.append(
                BlockIR(
                    kind=BlockKind.HEADING,
                    text=sheet.title.strip(),
                    metadata={
                        "chunk_boundary_candidate": True,
                        "chunk_boundary_reason": "sheet_name",
                        "chunk_boundary_confidence": 1.0,
                        "chunk_boundary_materialized": True,
                        "sheet": sheet.title,
                    },
                )
            )
        cells: list[TableCellIR] = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            for col_idx, value in enumerate(row):
                if value is None:
                    continue
                cells.append(
                    TableCellIR(
                        row_index=row_idx,
                        column_index=col_idx,
                        text=str(value),
                        is_header=row_idx == 0,
                    )
                )
        if cells:
            table_index += 1
            blocks.append(
                TableBlockIR(
                    cells=tuple(cells),
                    table_id=f"xlsx-table-{table_index}",
                    title=sheet.title,
                    metadata={"sheet": sheet.title},
                    merged_cells=bool(sheet.merged_cells.ranges),
                )
            )

    document = DocumentIR(source_format=DocumentFormat.XLSX, blocks=tuple(blocks), metadata={"sheet_count": len(workbook.worksheets)})
    return ParseResult(parser_id="openpyxl", document=document)


def _parse_doc_via_conversion(request: ParseRequest) -> ParseResult:
    with TemporaryDirectory(prefix="markbridge_doc_convert_") as tmp_dir:
        conversion = convert_doc_to_docx(Path(request.source_path), Path(tmp_dir))
        if not conversion.succeeded or conversion.output_path is None:
            raise ValueError(conversion.message or "DOC conversion failed.")
        converted_request = ParseRequest(
            source_path=conversion.output_path,
            document_format=DocumentFormat.DOCX,
            inspection=request.inspection,
            options=dict(request.options),
        )
        result = _parse_docx(converted_request)
        return ParseResult(
            parser_id="libreoffice",
            document=result.document,
            warnings=result.warnings + ((conversion.message,) if conversion.message else ()),
            metadata={"conversion_output_path": str(conversion.output_path)},
        )


def _parse_doc_with_antiword(request: ParseRequest) -> ParseResult:
    extraction = extract_doc_text_with_antiword(Path(request.source_path))
    if not extraction.succeeded or extraction.text is None:
        raise ValueError(extraction.message or "antiword text extraction failed.")

    blocks = _blocks_from_markdown(extraction.text)
    document = DocumentIR(
        source_format=DocumentFormat.DOC,
        blocks=tuple(blocks),
        metadata={
            "preferred_markdown": extraction.text,
            "source": "antiword",
            "extraction_mode": "text_fallback",
        },
    )
    return ParseResult(
        parser_id="antiword",
        document=document,
        warnings=((extraction.message,) if extraction.message else ()),
        metadata={"extraction_mode": "text_fallback"},
    )


def _parse_hwp_with_hwp5txt(request: ParseRequest) -> ParseResult:
    extraction = extract_hwp_text_with_hwp5txt(Path(request.source_path))
    if not extraction.succeeded or extraction.text is None:
        raise ValueError(extraction.message or "hwp5txt text extraction failed.")

    blocks = _blocks_from_markdown(extraction.text)
    document = DocumentIR(
        source_format=DocumentFormat.HWP,
        blocks=tuple(blocks),
        metadata={
            "preferred_markdown": extraction.text,
            "source": "hwp5txt",
            "extraction_mode": "text_route",
        },
    )
    return ParseResult(
        parser_id="hwp5txt",
        document=document,
        warnings=((extraction.message,) if extraction.message else ()),
        metadata={"extraction_mode": "text_route"},
    )


def _blocks_from_markdown(markdown: str, *, default_page_number: int | None = None) -> list[BlockIR]:
    blocks: list[BlockIR] = []
    paragraph_buffer: list[tuple[str, int]] = []
    table_buffer: list[tuple[str, int]] = []
    table_index = 0

    def flush_paragraph() -> None:
        if not paragraph_buffer:
            return
        line_numbers = [line_number for _, line_number in paragraph_buffer]
        text = "\n".join(line for line, _ in paragraph_buffer).strip()
        paragraph_buffer.clear()
        if text:
            blocks.append(
                BlockIR(
                    kind=BlockKind.PARAGRAPH,
                    text=text,
                    metadata=_markdown_block_metadata(line_numbers, default_page_number=default_page_number),
                )
            )

    def flush_table() -> None:
        nonlocal table_index
        if not table_buffer:
            return
        line_numbers = [line_number for _, line_number in table_buffer]
        rows = [_split_markdown_table_row(line) for line, _ in table_buffer]
        table_buffer.clear()
        if not rows:
            return
        table_index += 1
        cells: list[TableCellIR] = []
        data_row_index = 0
        row_lengths: list[int] = []
        for raw_row in rows:
            if _is_markdown_separator_row(raw_row):
                continue
            row_lengths.append(len(raw_row))
            for column_index, value in enumerate(raw_row):
                if value:
                    cells.append(
                        TableCellIR(
                            row_index=data_row_index,
                            column_index=column_index,
                            text=value,
                            is_header=data_row_index == 0,
                        )
                    )
            data_row_index += 1
        if cells:
            blocks.append(
                TableBlockIR(
                    cells=tuple(cells),
                    table_id=f"docling-table-{table_index}",
                    title=f"Table {table_index}",
                    merged_cells=len(set(row_lengths)) > 1,
                    metadata={
                        "source": "markdown_table",
                        "row_lengths": row_lengths,
                        **_markdown_block_metadata(line_numbers, default_page_number=default_page_number),
                    },
                )
            )

    for line_number, raw_line in enumerate(markdown.splitlines(), start=1):
        line = raw_line.rstrip()
        stripped = line.strip()
        if not stripped:
            flush_paragraph()
            flush_table()
            continue
        if stripped.startswith("|") and stripped.endswith("|"):
            flush_paragraph()
            table_buffer.append((stripped, line_number))
            continue
        if table_buffer:
            flush_table()
        if stripped.startswith("#"):
            flush_paragraph()
            level = len(stripped) - len(stripped.lstrip("#"))
            blocks.append(
                BlockIR(
                    kind=BlockKind.HEADING,
                    text=stripped[level:].strip(),
                    metadata={
                        "level": level,
                        "chunk_boundary_candidate": True,
                        "chunk_boundary_reason": "markdown_heading",
                        "chunk_boundary_confidence": 1.0,
                        "chunk_boundary_materialized": True,
                        **_markdown_block_metadata([line_number], default_page_number=default_page_number),
                    },
                )
            )
            continue
        if stripped.startswith(("- ", "* ")):
            flush_paragraph()
            blocks.append(
                BlockIR(
                    kind=BlockKind.LIST,
                    text=stripped[2:].strip(),
                    metadata=_markdown_block_metadata([line_number], default_page_number=default_page_number),
                )
            )
            continue
        paragraph_buffer.append((stripped, line_number))

    flush_paragraph()
    flush_table()
    return blocks


def _split_markdown_table_row(line: str) -> list[str]:
    trimmed = line.strip().strip("|")
    return [cell.strip() for cell in trimmed.split("|")]


def _is_markdown_separator_row(cells: list[str]) -> bool:
    return bool(cells) and all(cell and set(cell) <= {"-", ":"} for cell in cells)


def _markdown_block_metadata(line_numbers: list[int], *, default_page_number: int | None) -> dict[str, object]:
    metadata: dict[str, object] = {"markdown_line_numbers": line_numbers}
    if default_page_number is not None:
        metadata["page_number"] = default_page_number
    return metadata


_NUMBERED_HEADING_RE = re.compile(r"^(?:\d+\)|\d+(?:\.\d+){0,3}\.?|[IVXLCM]+\.[)]?|[A-Z]\.[)]?|[가-하]\.)\s+\S")
_CIRCLED_NUMBER_HEADING_RE = re.compile(r"^[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳]\s*\S")
_KOREAN_SECTION_RE = re.compile(r"^제\s*\d+\s*(?:장|절|조)\b")
_CLOSED_NUMBER_SECTION_RE = re.compile(r"^\d+\)\s+\S")
_SHORT_TITLE_KEYWORDS = (
    "보장내용",
    "보험금 지급",
    "가입나이",
    "보험기간",
    "보험료",
    "해약환급금",
    "지급사유",
    "지급금액",
    "계약내용",
    "보장명",
)
_SECTION_LABEL_KEYWORDS = (
    "대리인",
    "중요 확인 사항",
    "유선본인확인 업무",
    "비대면 접수 불가 대상건",
    "계약자 변경 구비서류",
    "계약자 변경 유의사항",
    "계약자 변경 가능범위",
    "동의 대상",
    "변경 가능기준",
    "공통사항",
    "구비서류",
)


def _next_nonempty_docx_item_hint(items: list[Paragraph | Table], *, start_index: int) -> tuple[str | None, str | None]:
    for item in items[start_index:]:
        if isinstance(item, Paragraph):
            text = item.text.strip()
            if text:
                return "paragraph", text
            continue
        if isinstance(item, Table):
            rows_text = _docx_table_rows_text(item)
            flattened = " ".join(value.strip() for row in rows_text for value in row if value and value.strip()).strip()
            return "table", flattened or None
    return None, None


def _adjacent_nonempty_docx_paragraph_text(
    items: list[Paragraph | Table], *, start_index: int, reverse: bool = False
) -> str | None:
    indices = range(start_index - 1, -1, -1) if reverse else range(start_index, len(items))
    for index in indices:
        item = items[index]
        if not isinstance(item, Paragraph):
            continue
        text = item.text.strip()
        if text:
            return text
    return None


def _has_nearby_circled_number_paragraph(
    items: list[Paragraph | Table], *, start_index: int, lookaround_limit: int = 3, reverse: bool = False
) -> bool:
    seen_paragraphs = 0
    indices = range(start_index - 1, -1, -1) if reverse else range(start_index, len(items))
    for index in indices:
        item = items[index]
        if not isinstance(item, Paragraph):
            continue
        text = item.text.strip()
        if not text:
            continue
        seen_paragraphs += 1
        if seen_paragraphs > lookaround_limit:
            return False
        if _CIRCLED_NUMBER_HEADING_RE.match(" ".join(text.split())):
            return True
    return False


def _iter_docx_block_items(parent: DocxPackageDocument) -> list[Paragraph | Table]:
    items: list[Paragraph | Table] = []
    for child in parent.element.body.iterchildren():
        if isinstance(child, CT_P):
            items.append(Paragraph(child, parent))
            continue
        if isinstance(child, CT_Tbl):
            items.append(Table(child, parent))
    return items


def _docx_heading_hint(
    *,
    text: str,
    style_name: str | None,
    next_text: str | None,
    next_item_kind: str | None,
    previous_paragraph_text: str | None = None,
    nearby_circled_sequence: bool = False,
) -> dict[str, object] | None:
    normalized = " ".join(text.split())
    if not normalized:
        return None

    style_hint = _docx_style_heading_hint(style_name)
    if style_hint is not None:
        return style_hint

    if len(normalized) > 120:
        return None

    if _NUMBERED_HEADING_RE.match(normalized):
        return {
            "reason": "heading_pattern.numbered",
            "confidence": 0.96,
            "level": _heading_level_from_numbered_text(normalized),
        }

    if _KOREAN_SECTION_RE.match(normalized):
        return {
            "reason": "heading_pattern.korean_section",
            "confidence": 0.98,
            "level": _heading_level_from_korean_section(normalized),
        }

    if _looks_like_circled_number_section_heading(
        normalized,
        next_text=next_text,
        next_item_kind=next_item_kind,
        previous_paragraph_text=previous_paragraph_text,
        nearby_circled_sequence=nearby_circled_sequence,
    ):
        return {"reason": "heading_pattern.circled_number_section", "confidence": 0.87, "level": 2}

    if _looks_like_closed_number_section_heading(normalized, next_text=next_text):
        return {"reason": "heading_pattern.list_section", "confidence": 0.84, "level": 2}

    if _looks_like_short_heading(normalized, next_text=next_text):
        return {"reason": "heading_pattern.short_title", "confidence": 0.72, "level": 2}

    return None


def _looks_like_short_heading(text: str, *, next_text: str | None) -> bool:
    if len(text) > 32:
        return False
    if any(punctuation in text for punctuation in (". ", "? ", "! ", "다.", "요.")):
        return False
    if text.count(" ") > 4:
        return False
    if next_text is None or len(next_text.strip()) < 20:
        return False
    if any(keyword in text for keyword in _SHORT_TITLE_KEYWORDS):
        return True
    if ":" in text or " - " in text:
        return False
    return text.endswith(("사항", "기준", "내용", "조건", "정의"))


def _looks_like_closed_number_section_heading(text: str, *, next_text: str | None) -> bool:
    if not _CLOSED_NUMBER_SECTION_RE.match(text):
        return False
    label = _closed_number_section_label(text)
    if not label or len(label) > 24:
        return False
    if any(punctuation in label for punctuation in (". ", "? ", "! ", "다.", "요.", ":", " - ")):
        return False
    if not any(keyword in label for keyword in _SECTION_LABEL_KEYWORDS):
        return False
    if next_text is None or len(next_text.strip()) < 12:
        return False
    return True


def _looks_like_circled_number_section_heading(
    text: str,
    *,
    next_text: str | None,
    next_item_kind: str | None,
    previous_paragraph_text: str | None,
    nearby_circled_sequence: bool,
) -> bool:
    if not _CIRCLED_NUMBER_HEADING_RE.match(text):
        return False
    if len(text) > 48:
        return False
    if any(punctuation in text for punctuation in ("다.", "요.", ": ", " - ")):
        return False
    if next_item_kind == "table":
        return True
    if previous_paragraph_text and _CIRCLED_NUMBER_HEADING_RE.match(" ".join(previous_paragraph_text.split())):
        return True
    if next_text and _CIRCLED_NUMBER_HEADING_RE.match(" ".join(next_text.split())):
        return True
    if nearby_circled_sequence:
        return True
    if next_text is None or len(next_text.strip()) < 20:
        return False
    return text.endswith(("조회", "납입", "신청", "변경", "관리", "발급", "서비스", "지원", "업무"))


def _docx_style_heading_hint(style_name: str | None) -> dict[str, object] | None:
    if not style_name:
        return None
    normalized_style = style_name.strip().lower()
    if "heading" in normalized_style:
        level = _heading_level_from_style_name(style_name.strip())
        return {"reason": "heading_style", "confidence": 1.0, "level": level}
    korean_style = style_name.strip()
    if "제목" in korean_style:
        level = _heading_level_from_style_name(korean_style)
        return {"reason": "heading_style.custom_title", "confidence": 0.99, "level": level}
    if "목차" in korean_style or "장제목" in korean_style:
        return {"reason": "heading_style.custom_title", "confidence": 0.99, "level": 2}
    return None


def _docx_table_rows_text(table: object) -> list[list[str]]:
    rows_text: list[list[str]] = []
    for row in getattr(table, "rows", ()):
        values: list[str] = []
        seen: set[str] = set()
        for cell in getattr(row, "cells", ()):
            text = getattr(cell, "text", "").strip()
            if not text:
                continue
            if text in seen:
                continue
            seen.add(text)
            values.append(text)
        rows_text.append(values)
    return rows_text


def _normalize_docx_table_rows(table: Table) -> tuple[list[list[str]], bool]:
    raw_rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
    if not raw_rows:
        return [], False
    width = max(len(row) for row in raw_rows)
    normalized = [row + ([""] * (width - len(row))) for row in raw_rows]
    normalized, suppressed_horizontal_duplicates = _suppress_docx_horizontal_merge_duplicates(normalized)
    applied_carry_forward = False

    for row_index in range(1, len(normalized)):
        row = normalized[row_index]
        if not any(value.strip() for value in row):
            continue
        for column_index in range(min(2, width)):
            current = row[column_index].strip()
            if current:
                continue
            previous = normalized[row_index - 1][column_index].strip()
            if _should_carry_forward_docx_cell(
                previous_value=previous,
                row=row,
                column_index=column_index,
            ):
                row[column_index] = previous
                applied_carry_forward = True

    normalized, removed_empty_columns = _drop_empty_docx_columns(normalized)

    return normalized, applied_carry_forward or suppressed_horizontal_duplicates or removed_empty_columns


def _suppress_docx_horizontal_merge_duplicates(rows: list[list[str]]) -> tuple[list[list[str]], bool]:
    normalized = [row[:] for row in rows]
    changed = False

    for row in normalized:
        for column_index in range(1, len(row)):
            current = row[column_index].strip()
            previous = row[column_index - 1].strip()
            if not current or current != previous:
                continue
            if not any(value.strip() for value in row[column_index + 1:]):
                continue
            row[column_index] = ""
            changed = True

    return normalized, changed


def _drop_empty_docx_columns(rows: list[list[str]]) -> tuple[list[list[str]], bool]:
    if not rows or not rows[0]:
        return rows, False

    keep_indices = [
        column_index
        for column_index in range(len(rows[0]))
        if any(column_index < len(row) and row[column_index].strip() for row in rows)
    ]
    if len(keep_indices) == len(rows[0]):
        return rows, False

    collapsed = [[row[column_index] for column_index in keep_indices] for row in rows]
    return collapsed, True


def _should_carry_forward_docx_cell(*, previous_value: str, row: list[str], column_index: int) -> bool:
    if not previous_value:
        return False
    if "\n" in previous_value or len(previous_value) > 30:
        return False
    if column_index > 1:
        return False
    later_values = [value.strip() for value in row[column_index + 1:]]
    if not any(later_values):
        return False
    if any(value == previous_value for value in later_values[:1]):
        return False
    return True


def _is_docx_layout_table(rows_text: list[list[str]]) -> bool:
    nonempty_rows = [row for row in rows_text if row]
    if not nonempty_rows:
        return False
    if any(len(row) > 1 for row in nonempty_rows):
        return False
    return True


def _blocks_from_docx_layout_rows(rows_text: list[list[str]]) -> list[BlockIR]:
    values = [value.strip() for row in rows_text if row for value in row if value and value.strip()]
    if not values:
        return []
    note_text = "\n\n".join(values)
    return [BlockIR(kind=BlockKind.NOTE, text=note_text, metadata={"source": "docx_layout_table", "box_preserved": True})]


def _blocks_from_docx_layout_text(text: str) -> list[BlockIR]:
    blocks: list[BlockIR] = []
    paragraph_buffer: list[str] = []

    def flush_paragraph() -> None:
        if not paragraph_buffer:
            return
        paragraph = "\n".join(paragraph_buffer).strip()
        paragraph_buffer.clear()
        if paragraph:
            blocks.append(BlockIR(kind=BlockKind.PARAGRAPH, text=paragraph, metadata={"source": "docx_layout_table"}))

    lines = [line.strip() for line in text.splitlines()]
    for index, line in enumerate(lines):
        if not line:
            flush_paragraph()
            continue
        heading_hint = _layout_heading_hint(line)
        if heading_hint is not None:
            flush_paragraph()
            blocks.append(
                BlockIR(
                    kind=BlockKind.HEADING,
                    text=line,
                    metadata={
                        "source": "docx_layout_table",
                        "level": heading_hint["level"],
                        "chunk_boundary_candidate": True,
                        "chunk_boundary_reason": heading_hint["reason"],
                        "chunk_boundary_confidence": heading_hint["confidence"],
                        "chunk_boundary_materialized": True,
                    },
                )
            )
            continue
        if line.startswith(("■", "※", "-", "*")):
            flush_paragraph()
            blocks.append(BlockIR(kind=BlockKind.LIST, text=line, metadata={"source": "docx_layout_table"}))
            continue
        if _is_line_continuation_candidate(line) and paragraph_buffer:
            paragraph_buffer.append(line)
            continue
        if index > 0 and paragraph_buffer and _looks_like_new_statement(line):
            flush_paragraph()
        paragraph_buffer.append(line)

    flush_paragraph()
    return blocks


def _layout_heading_hint(text: str) -> dict[str, object] | None:
    normalized = " ".join(text.split())
    if not normalized:
        return None
    if _NUMBERED_HEADING_RE.match(normalized) and len(normalized) <= 80:
        return {
            "reason": "heading_pattern.numbered",
            "confidence": 0.9,
            "level": _heading_level_from_numbered_text(normalized),
        }
    if _KOREAN_SECTION_RE.match(normalized):
        return {
            "reason": "heading_pattern.korean_section",
            "confidence": 0.95,
            "level": _heading_level_from_korean_section(normalized),
        }
    return None


def _is_line_continuation_candidate(text: str) -> bool:
    return text.startswith(("→", "(", "[", "ⓐ", "ⓑ", "①", "②", "③", "④", "⑤", "⑥", "⑦", "⑧", "⑨", "⑩"))


def _looks_like_new_statement(text: str) -> bool:
    return len(text) <= 120 and not text.startswith(("→", "(", "["))


def _heading_level_from_style_name(style_name: str) -> int:
    normalized = style_name.strip().lower()
    if "heading 1" in normalized or "제목 1" in normalized or "첫제목" in style_name:
        return 2
    if "heading 2" in normalized or "제목 2" in normalized or "두번째제목" in style_name:
        return 3
    if "heading 3" in normalized or "제목 3" in normalized or "세번째제목" in style_name:
        return 4
    return 2


def _heading_level_from_numbered_text(text: str) -> int:
    leading = text.split(maxsplit=1)[0]
    digit_part = leading.rstrip(".")
    if re.fullmatch(r"\d+(?:\.\d+)+", digit_part):
        depth = digit_part.count(".") + 1
        return min(5, depth + 1)
    return 2


def _heading_level_from_korean_section(text: str) -> int:
    if "장" in text:
        return 2
    if "절" in text:
        return 3
    if "조" in text:
        return 4
    return 2


def _closed_number_section_label(text: str) -> str:
    parts = text.split(")", 1)
    if len(parts) != 2:
        return ""
    return parts[1].strip()
